import sqlite3
import os
import bcrypt

from openpyxl.styles import Font, Border, Side, Alignment

from ExcelBuilder import ExcelBuilder
from config import categories, faculties

# Класс Базы Данных
class DataBase:
    def __init__(self):
        try:
            database_path = os.path.join(os.path.dirname(__file__), 'data', "qms_database.db")
            db_exists = os.path.exists(database_path)

            self.connection = sqlite3.connect(database_path)
            self.cursor = self.connection.cursor()

            # Создание базы данных, если не существует
            if not db_exists:
                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS admins (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    login TEXT NOT NULL UNIQUE,
                    password_hash BLOB NOT NULL
                );
                ''')

                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL
                );
                ''')

                self.cursor.execute('''
                    INSERT INTO categories (name)
                    VALUES (?), (?), (?);
                ''', categories)

                self.cursor.execute('''SELECT name, id FROM categories''')
                form_id = dict(self.cursor.fetchall())

                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS questions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    
                    category_id INTEGER NOT NULL,
                    
                    number INTEGER NOT NULL,
                    text TEXT NOT NULL,
                    type BOOLEAN NOT NULL,
                    
                    FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
                );
                ''')

                questions = {"Студент":
                                 [["Уровень теоретической подготовки",
                                   "Уровень практической подготовки",
                                   "Уровень и возможность освоения "
                                       "дополнительных знаний и умений: "
                                       "владение современными информационными "
                                       "технологиями, владение иностранными "
                                       "языками и др.",
                                   "Обеспеченность образовательного "
                                       "процесса учебными и учебно-"
                                       "методическими пособиями",
                                   "Доступность в Научной библиотеке "
                                        "БНТУ электронных ресурсов по тематике обучения",
                                   "Обеспеченность образовательного "
                                       "процесса учебной материально-"
                                       "технической базой (оборудованием)",
                                   "Уровень организации контроля и оценки знаний",
                                   "Содержание и организация общественных, "
                                   "культурно-массовых, спортивных и др мероприятий"],
                                  ["Пожалуйста, перечислите Ф.И.О, преподавателей (не более 3) с указанием "
                                       "преподаваемых дисциплин, работой которых Вы НАИБОЛЕЕ удовлетворены в текущем "
                                       "учебном году",
                                   "Пожалуйста, выскажите ВАШИ предложения и замечания относительно содержания, "
                                        "организации и обеспечения образовательного процесса"]],
                             "Выпускник":
                                 [["Уровень теоретической подготовки",
                                   "Уровень практической подготовки",
                                   "Уровень освоения дополнительных"
                                        "знаний и умений: владение современными "
                                        "информационными технологиями, владение "
                                        "иностранными языками и др.",
                                   "Владение навыками самостоятельной работы",
                                   "Готовность к профессиональной деятельности"],
                                  ["Пожалуйста, выскажите ВАШИ предложения и замечания относительно содержания, "
                                   "организации и обеспечения образовательного процесса в Университете"]],
                             "Магистрант":
                                 [["Уровень изучения учебных дисциплин",
                                   "Уровень организации самостоятельной работы магистрантов",
                                   "Взаимодействие с научным руководителем",
                                   "Возможность взаимодействия с внешними организациями по тематике исследования",
                                   "Контроль со стороны кафедры за процессом обучения в магистратуре",
                                   "Возможность представления результатов НИР на конференциях и семинарах",
                                   "Обеспеченность образовательного процесса учебными и учебно-методическими пособиями",
                                   "Доступность в Научной библиотеке БНТУ научных "
                                        "электронных ресурсов по тематике исследования",
                                   "Уровень материально-технического обеспечения учебного процесса и НИР"],
                                  ["Пожалуйста, выскажите ВАШИ предложения и замечания относительно содержания, "
                                   "организации и обеспечения учебного процесса, а также НИР магистрантов "
                                   "в Университете"]]
                             }

                for category in questions.keys():

                    for question_type in range(2):
                        question_number = question_type + 1

                        for question in questions[category][question_type]:
                            self.cursor.execute('''
                                            INSERT INTO questions (category_id,
                                                                    number,
                                                                    text, 
                                                                    type)
                                            VALUES (?, ?, ?, ?);
                                        ''', (form_id[category], question_number, question, question_type))
                            question_number += 1


                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS faculties (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE
                );
                ''')

                for faculty in faculties:
                    self.cursor.execute('''
                                    INSERT INTO faculties (name)
                                    VALUES (?);
                                ''', (faculty,))


                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS dates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    
                    faculty_id INTEGER NOT NULL,
                    category_id INTEGER NOT NULL,
                                        
                    date TEXT NOT NULL,
                    course INTEGER,
                    
                    FOREIGN KEY (faculty_id) REFERENCES faculties(id) ON DELETE CASCADE,
                    FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
                );
                ''')

                self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS answers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    
                    date_id INTEGER NOT NULL,
                    question_id INTEGER NOT NULL,
                    
                    answer TEXT NOT NULL,
                    amount INTEGER NOT NULL,
                    
                    FOREIGN KEY (date_id) REFERENCES dates(id) ON DELETE CASCADE,
                    FOREIGN KEY (question_id) REFERENCES questions(id) ON DELETE CASCADE
                );
                ''')

            self.connection.commit()

            print("База Данных подключена")

        except Exception as ex:
            print(f"ERROR: {ex}")

    # Проверка данных входа
    def check_enter_data(self, login, password):
        send_data = {"access": False}

        self.cursor.execute('''
            SELECT password_hash FROM admins
            WHERE login = ?
            ''', (login,))

        found_hash = self.cursor.fetchone()

        if found_hash is not None:
            if bcrypt.checkpw(password.encode('utf-8'), found_hash[0]):
                send_data["access"] = True
            else:
                send_data["error"] = "Неверный пароль"
        else:
            send_data["error"] = "Пользователя с указанным логином не существует"

        return send_data

    # Добавление учётной записи
    def new_admin(self, login, password):
        send_data = {"access": False}

        self.cursor.execute('''
            SELECT * FROM admins
            WHERE login = ?
            ''', (login,))

        if self.cursor.fetchone() is None:
            self.cursor.execute('INSERT INTO admins (login, password_hash) VALUES (?, ?)',
                                (login, bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())))

            self.connection.commit()
            send_data["access"] = True
        else:
            send_data["error"] = "Пользователь с введённым логином уже существует"

        return send_data

    # Получение вопросов
    def get_questions(self, category):
        self.cursor.execute('''SELECT name, id FROM categories''')
        form_id = dict(self.cursor.fetchall())

        questions = []

        for quest_type in range(2):
            self.cursor.execute('''
                        SELECT number, text, id FROM questions
                        WHERE category_id = ?
                        AND type = ?
                        ''', (form_id[category], quest_type))

            questions.append({key: (q_id, value) for key, value, q_id in self.cursor.fetchall()})

        return questions

    # Запись ответов
    def write_answers(self, data):
        self.cursor.execute('''
                        SELECT id FROM faculties
                        WHERE name = ?
                        ''', (data["info"]["faculty"],))
        faculty_id = self.cursor.fetchone()[0]

        self.cursor.execute('''
                                SELECT id FROM categories
                                WHERE name = ?
                                ''', (data["info"]["category"],))
        category_id = self.cursor.fetchone()[0]

        course = 0 if data["info"]["course"] is None else data["info"]["course"]

        while True:
            self.cursor.execute('SELECT id FROM dates WHERE '
                                'faculty_id = ? AND category_id = ? AND course = ?',
                                (faculty_id, category_id, course))

            found_date = self.cursor.fetchone()
            if found_date is None:
                self.cursor.execute('INSERT INTO dates (faculty_id, category_id, date, course) VALUES (?, ?, ?, ?)',
                                (faculty_id, category_id, data["info"]["date"], course))
            else:
                date_id = found_date[0]
                break

        for question_id, answer_number in data["answers"][0].items():
            self.cursor.execute('SELECT id, amount FROM answers WHERE '
                                'date_id = ? AND question_id = ? AND answer = ?',
                                (date_id, question_id, answer_number))

            found_answer = self.cursor.fetchone()
            if found_answer is None:
                self.cursor.execute('INSERT INTO answers (date_id, question_id, answer, amount) '
                                    'VALUES (?, ?, ?, ?)',
                                    (date_id, question_id, answer_number, 1))
            else:
                answer_id, amount = found_answer
                self.cursor.execute("UPDATE answers SET amount = ? WHERE id = ?", (amount + 1, answer_id))

        for question_id, answer in data["answers"][1].items():
            self.cursor.execute('INSERT INTO answers (date_id, question_id, answer, amount) '
                                'VALUES (?, ?, ?, ?)',
                                (date_id, question_id, answer, 1))

        self.connection.commit()


    # Получение ответов
    def get_answers(self, data):
        self.cursor.execute('''
                        SELECT id FROM faculties
                        WHERE name = ?
                        ''', (data["faculty"],))
        faculty_id = self.cursor.fetchone()[0]

        self.cursor.execute('''
                                SELECT id FROM categories
                                WHERE name = ?
                                ''', (data["category"],))
        category_id = self.cursor.fetchone()[0]

        course = 0 if data["course"] is None else data["course"]

        self.cursor.execute('SELECT id FROM dates WHERE '
                                'faculty_id = ? AND category_id = ? AND course = ?',
                                (faculty_id, category_id, course))

        found_date = self.cursor.fetchone()

        if found_date is None:
            return None

        date_id = found_date[0]

        self.cursor.execute('SELECT question_id, answer, amount FROM answers WHERE '
                                'date_id = ?',
                                (date_id,))

        answers_list = self.cursor.fetchall()

        send_data = {"answers": [{}, {}], "averages": [0,0,0,0]}

        for question_id, answer, amount in answers_list:
            self.cursor.execute('SELECT number, text, type FROM questions WHERE '
                                'id = ?',
                                (question_id,))
            number, text, type = self.cursor.fetchone()
            if number not in send_data["answers"][type].keys():
                send_data["answers"][type][number] = {"text": text, "answers": ({'1': 0, '2': 0, '3': 0, '4': 0} if type == 0 else [])}

            if type == 0:
                send_data["answers"][type][number]["answers"][answer] = amount
            else:
                send_data["answers"][type][number]["answers"].append(answer)

        sum_amount = sum(send_data["answers"][0][1]["answers"].values())
        for quest_number, descr in send_data["answers"][0].items():
            for answer, amount in descr['answers'].items():
                send_data["averages"][int(answer)-1] += amount
            sum_persent = descr['answers']['1'] + descr['answers']['2'] / 2
            send_data["answers"][0][quest_number]['persent'] = sum_persent/sum_amount * 100

        send_data["amount"] = sum_amount

        send_data["averages"] = [summ/len(send_data["answers"][0]) for summ in send_data["averages"]]

        send_data["ultimate_satisfaction"] = (send_data["averages"][0] + send_data["averages"][1] / 2) / sum_amount * 100

        return send_data

    # Получение файла отчёта Excel
    def get_faculty_stat(self, faculty):

        titles = {"Студент": "Студенты",
                  "Выпускник": "Выпускники",
                  "Магистрант": "Магистранты"}

        excel_b = ExcelBuilder(faculty)
        row = excel_b.row
        sheet = excel_b.sheet

        avar1 = []
        avar2 = []
        amounts = []
        percents = []

        self.cursor.execute('''
                                SELECT id FROM faculties
                                WHERE name = ?
                                ''', (faculty,))
        faculty_id = self.cursor.fetchone()[0]

        exist = False

        for category in titles.keys():
            self.cursor.execute('''
                                       SELECT id FROM categories
                                       WHERE name = ?
                                       ''', (category,))
            category_id = self.cursor.fetchone()[0]

            courses = [0]

            if category == 'Студент':

                self.cursor.execute('''
                                        SELECT course FROM dates WHERE
                                        faculty_id = ? AND category_id = ?''',
                                    (faculty_id, category_id))

                courses = list(set(el[0] for el in self.cursor.fetchall()))

            for course in courses:
                if category == 'Студент' and courses[0] == course:
                    sheet.merge_cells(f'A{row}:D{row}')
                    sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
                    sheet.cell(row, 1).value = titles[category]
                    sheet.cell(row, 1).font = Font(size=16)

                    row += 2

                if course != 0:
                    sheet.merge_cells(f'A{row}:D{row}')
                    sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
                    sheet.cell(row, 1).value = f"{course} курс"
                    sheet.cell(row, 1).font = Font(size=16)

                self.cursor.execute('''
                                    SELECT id FROM dates WHERE
                                    faculty_id = ? AND category_id = ? AND course = ?''',
                                    (faculty_id, category_id, course))

                found_date = self.cursor.fetchone()

                if found_date is not None:
                    exist = True
                    date_id = found_date[0]

                    if category != 'Студент':
                        sheet.merge_cells(f'A{row}:D{row}')
                        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
                        sheet.cell(row, 1).value = titles[category]
                        sheet.cell(row, 1).font = Font(size=16)

                    row += 2

                    sheet.cell(row, 1).value = "№ вопроса"
                    sheet.cell(row, 2).value = "Удовлетворён"
                    sheet.cell(row, 3).value = "Скорее удовлетворён"
                    sheet.cell(row, 4).value = "УП i"
                    for col in range(1, 5):
                        sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                        sheet.cell(row, col).border = Border(
                            left=Side(style='thin'),
                            top=Side(style='thin'),
                            right=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    row += 1

                    self.cursor.execute('''
                                            SELECT question_id, answer, amount FROM answers WHERE
                                            date_id = ? ''',
                                        (date_id,))
                    answers = self.cursor.fetchall()

                    answers_dict = {}

                    for question_id, answer, amount in answers:
                        self.cursor.execute('SELECT number, type FROM questions WHERE '
                                            'id = ?',
                                            (question_id,))
                        number, type = self.cursor.fetchone()
                        if not type:
                            if number not in answers_dict.keys():
                                answers_dict[number] = {'1': 0, '2': 0, '3': 0, '4': 0}

                            answers_dict[number][answer] = amount

                    row_sum = row + len(answers_dict) - 1

                    amount_val = sum(answers_dict[1].values())
                    sheet.cell(row_sum, 5).value = amount_val

                    sheet.cell(row_sum, 5).alignment = Alignment(horizontal='center', vertical='center')
                    sheet.cell(row_sum, 5).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    curr_row = row + len(answers_dict)

                    amounts.append(curr_row)

                    sheet.cell(curr_row, 5).value = f"=(B{curr_row}+(C{curr_row}/2))/E{row_sum}"
                    sheet.cell(curr_row, 5).number_format = '0.00%'

                    sheet.cell(curr_row, 5).alignment = Alignment(horizontal='center', vertical='center')
                    sheet.cell(curr_row, 5).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    avar1.append(f"B{row}:B{row + len(answers_dict) - 1}")
                    avar2.append(f"C{row}:C{row + len(answers_dict) - 1}")

                    summ1 = 0
                    summ2 = 0

                    for number, answers in answers_dict.items():
                        sheet.cell(row, 1).value = number
                        sheet.cell(row, 2).value = answers['1']
                        summ1 += answers['1']
                        sheet.cell(row, 3).value = answers['2']
                        summ2 += answers['2']
                        sheet.cell(row, 4).value = f"=(B{row}+C{row}/2)/$E${row_sum}"
                        sheet.cell(row, 4).number_format = '0.00%'
                        for col in range(1, 5):
                            sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                            sheet.cell(row, col).border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                        row += 1

                    sheet.cell(row, 1).value = "Ср"
                    sheet.cell(row, 2).value = f"=SUM(B{row - len(answers_dict)}:B{row - 1})/{len(answers_dict)}"
                    avar_var_1 = summ1/len(answers_dict)
                    sheet.cell(row, 2).number_format = '0.###'
                    sheet.cell(row, 3).value = f"=SUM(C{row - len(answers_dict)}:C{row - 1})/{len(answers_dict)}"
                    avar_var_2 = summ2/len(answers_dict)
                    sheet.cell(row, 3).number_format = '0.###'
                    sheet.cell(row, 4).value = "УП"
                    for col in range(1, 5):
                        sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
                        sheet.cell(row, col).border = Border(
                            left=Side(style='thin'),
                            top=Side(style='thin'),
                            right=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                    percents.append((amount_val, (avar_var_1 + avar_var_2/2)/amount_val))

                    row += 3

        if not exist:
            return None

        return excel_b.build_end(amounts, percents, avar1, avar2)

    # 'info': {'category': 'Студент', 'faculty': 'ФИТР', 'course': 3, 'date': '16.05.2025'}
    def insert_answers(self, info, amounts):
        self.cursor.execute('''
                                SELECT id FROM faculties
                                WHERE name = ?
                                ''', (info["faculty"],))
        faculty_id = self.cursor.fetchone()[0]

        self.cursor.execute('''
                                        SELECT id FROM categories
                                        WHERE name = ?
                                        ''', (info["category"],))
        category_id = self.cursor.fetchone()[0]

        course = 0 if info["course"] is None else info["course"]

        while True:
            self.cursor.execute('SELECT id FROM dates WHERE '
                                'faculty_id = ? AND category_id = ? AND course = ?',
                                (faculty_id, category_id, course))

            found_date = self.cursor.fetchone()
            if found_date is None:
                self.cursor.execute('INSERT INTO dates (faculty_id, category_id, date, course) VALUES (?, ?, ?, ?)',
                                    (faculty_id, category_id, info["date"], course))
            else:
                date_id = found_date[0]
                break

        for number in range(len(amounts)):
            self.cursor.execute('SELECT id FROM questions WHERE '
                                'category_id = ? AND number = ? AND type = ?',
                                (category_id, number, False))

            question_id = self.cursor.fetchone()[0]

            for answer_number in range(4):
                amount = amounts[number][answer_number]

                self.cursor.execute('SELECT id FROM answers WHERE '
                                    'date_id = ? AND question_id = ? AND answer = ?',
                                    (date_id, question_id, answer_number))

                found_answer = self.cursor.fetchone()
                if found_answer is None and amount != 0:
                    self.cursor.execute('INSERT INTO answers (date_id, question_id, answer, amount) '
                                        'VALUES (?, ?, ?, ?)',
                                        (date_id, question_id, answer_number, amounts[number][answer_number]))
                else:
                    answer_id = found_answer[0]
                    if amount == 0:
                        self.cursor.execute("DELETE FROM answers WHERE id = answer_id;")
                    else:
                        self.cursor.execute("UPDATE answers SET amount = ? WHERE id = ?", (amount, answer_id))

        self.connection.commit()