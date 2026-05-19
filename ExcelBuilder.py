import base64
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border

class ExcelBuilder:
    def __init__(self, faculty):
        self.wb = openpyxl.Workbook()

        self.sheet = self.wb.active

        self.sheet.title = faculty

        self.sheet.column_dimensions['A'].width = 10
        self.sheet.column_dimensions['B'].width = 16
        self.sheet.column_dimensions['C'].width = 22
        self.sheet.column_dimensions['D'].width = 9
        self.sheet.column_dimensions['H'].width = 16
        self.sheet.column_dimensions['I'].width = 22

        self.sheet.cell(1, 1).value = f"Итоги анкетирования {faculty}"
        self.sheet.cell(1, 1).font = Font(size=24)

        self.row = 3
        
    def build_end(self, amounts, percents, avar1, avar2):
        self.sheet.cell(8, 7).value = "Ср"

        self.sheet.cell(7, 8).value = "Удовлетворён"
        self.sheet.cell(7, 8).alignment = Alignment(horizontal='center', vertical='center')
        self.sheet.cell(8, 8).value = f"=AVERAGE({','.join(avar1)})"
        self.sheet.cell(8, 8).number_format = '0.##'

        self.sheet.cell(7, 9).value = "Скорее удовлетворён"
        self.sheet.cell(7, 9).alignment = Alignment(horizontal='center', vertical='center')
        self.sheet.cell(8, 9).value = f"=AVERAGE({','.join(avar2)})"
        self.sheet.cell(8, 9).number_format = '0.##'

        self.sheet.cell(8, 10).value = "УП"
        self.sheet.cell(8, 11).value = '=' + '+'.join([f"E{row}/K7*E{row - 1}" for row in amounts])
        self.sheet.cell(8, 11).number_format = '0.00%'

        self.sheet.cell(7, 11).value = f"=SUM({','.join([f'E{row - 1}' for row in amounts])})"
        self.sheet.cell(7, 11).alignment = Alignment(horizontal='center', vertical='center')
        self.sheet.cell(7, 11).border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(7, 12):
            self.sheet.cell(8, col).alignment = Alignment(horizontal='center', vertical='center')
            self.sheet.cell(8, col).border = Border(
                left=Side(style='thin'),
                top=Side(style='thin'),
                right=Side(style='thin'),
                bottom=Side(style='thin')
            )

        buffer = BytesIO()
        self.wb.save(buffer)
        excel_bytes = buffer.getvalue()

        result_amount = sum([item[0] for item in percents])

        ultimate_satisfaction = sum([item[1] / result_amount * item[0] for item in percents]) * 100

        return base64.b64encode(excel_bytes).decode('utf-8'), ultimate_satisfaction